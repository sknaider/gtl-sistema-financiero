from fastapi import APIRouter, UploadFile, File, HTTPException, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
import pandas as pd
import io
from datetime import datetime
from difflib import SequenceMatcher
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill, Font, Alignment
import re
from core.database import get_db
from models.import_history import ImportHistory

router = APIRouter()

COLUMN_MAPPINGS = {
    'FECHA': ['date', 'dia', 'day', 'when', 'fecha_operacion', 'fecha_ingreso'],
    'CLIENTE': ['customer', 'company', 'compania', 'client', 'empresa', 'razon_social', 'name', 'nombre'],
    'MONTO': ['amount', 'value', 'valor', 'price', 'precio', 'total', 'importe'],
    'MONEDA': ['currency', 'coin', 'tipo_moneda', 'curr'],
    'AWB': ['waybill', 'guia', 'tracking', 'numero_guia', 'air_waybill'],
    'DESCRIPCION': ['description', 'detail', 'detalle', 'concepto', 'notes', 'notas'],
    'MES': ['month', 'periodo', 'period', 'mes_operacion']
}

def fuzzy_match_column(col_name: str, threshold: float = 0.7):
    col_lower = col_name.lower().strip()
    
    for target, aliases in COLUMN_MAPPINGS.items():
        if col_lower == target.lower():
            return target, 1.0
        for alias in aliases:
            if col_lower == alias.lower():
                return target, 0.95
    
    best_match = None
    best_score = 0
    
    for target, aliases in COLUMN_MAPPINGS.items():
        score = SequenceMatcher(None, col_lower, target.lower()).ratio()
        if score > best_score and score >= threshold:
            best_score = score
            best_match = target
        
        for alias in aliases:
            score = SequenceMatcher(None, col_lower, alias.lower()).ratio()
            if score > best_score and score >= threshold:
                best_score = score
                best_match = target
    
    return best_match, best_score if best_match else None

def clean_monto(value):
    """Limpia y convierte monto a float"""
    if pd.isna(value) or value == '' or value is None:
        return None
    
    val_str = str(value).strip()
    
    # Si está vacío después de strip
    if not val_str or val_str.lower() in ['nan', 'none', 'null', '#n/a', '-']:
        return None
    
    # Quitar todo excepto números, puntos y guión
    val_str = re.sub(r'[^\d.-]', '', val_str)
    
    if not val_str or val_str == '-':
        return None
    
    try:
        return float(val_str)
    except:
        return None

def parse_fecha(value):
    """Intenta parsear fecha en múltiples formatos"""
    if pd.isna(value) or value == '' or value is None:
        return None
    
    val_str = str(value).strip()
    
    # Valores vacíos o nulos
    if not val_str or val_str.lower() in ['nan', 'none', 'null', '#n/a', '-']:
        return None
    
    # Si ya es datetime
    if isinstance(value, (pd.Timestamp, datetime)):
        return value
    
    # Formatos comunes
    formatos = [
        '%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d', '%Y/%m/%d',
        '%d.%m.%Y', '%Y.%m.%d', '%d %m %Y', '%Y %m %d',
        '%d/%m/%y', '%d-%m-%y', '%y-%m-%d', '%y/%m/%d'
    ]
    
    for fmt in formatos:
        try:
            return datetime.strptime(val_str, fmt)
        except:
            continue
    
    # Último intento con pandas
    try:
        return pd.to_datetime(val_str, dayfirst=True)
    except:
        return None

@router.post("/excel/sheets")
async def get_excel_sheets(file: UploadFile = File(...)):
    """Obtener lista de hojas del Excel"""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(400, "Solo archivos Excel (.xlsx, .xls)")
    
    contents = await file.read()
    xl = pd.ExcelFile(io.BytesIO(contents))
    
    return {
        "sheets": xl.sheet_names,
        "filename": file.filename
    }

@router.post("/excel/preview")
async def preview_excel(file: UploadFile = File(...), sheet_name: str = None):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(400, "Solo archivos Excel (.xlsx, .xls)")
    
    contents = await file.read()
    
    # Si no se especifica hoja, usar la primera
    if sheet_name is None:
        xl = pd.ExcelFile(io.BytesIO(contents))
        sheet_name = xl.sheet_names[0]
    
    df = pd.read_excel(io.BytesIO(contents), sheet_name=sheet_name)
    
    is_vertical = df.iloc[0, 0] in ['FECHA', 'CLIENTE', 'MONTO', 'fecha', 'cliente', 'monto']
    
    if is_vertical:
        df = df.T
        df.columns = df.iloc[0]
        df = df[1:]
    
    column_mapping = {}
    unmapped_columns = []
    mapping_confidence = {}
    
    for col in df.columns:
        matched, confidence = fuzzy_match_column(str(col))
        if matched:
            column_mapping[col] = matched
            mapping_confidence[col] = confidence
        else:
            unmapped_columns.append(col)
    
    df_mapped = df.rename(columns=column_mapping)
    
    errores = []
    sugerencias = []
    
    required_cols = ['FECHA', 'CLIENTE', 'MONTO']
    missing = [col for col in required_cols if col not in df_mapped.columns]
    if missing:
        errores.append(f"⚠️ Columnas requeridas: {', '.join(missing)}")
        if unmapped_columns:
            sugerencias.append(f"💡 Columnas sin mapear: {', '.join(unmapped_columns)}")
    
    # LIMPIAR Y VALIDAR FECHA
    fechas_invalidas = []
    if 'FECHA' in df_mapped.columns:
        fechas_limpias = []
        
        for idx, val in df_mapped['FECHA'].items():
            fecha_limpia = parse_fecha(val)
            fechas_limpias.append(fecha_limpia)
            if fecha_limpia is None and pd.notna(val) and str(val).strip() != '':
                fechas_invalidas.append(idx + 2)
        
        df_mapped['FECHA'] = fechas_limpias
        
        if fechas_invalidas[:3]:
            errores.append(f"⚠️ Filas {fechas_invalidas[:3]} tienen fechas inválidas")
            sugerencias.append(f"💡 Puedes importar omitiendo las {len(fechas_invalidas)} filas con errores")
    
    # LIMPIAR Y VALIDAR MONTO
    montos_invalidos = []
    if 'MONTO' in df_mapped.columns:
        montos_limpios = []
        
        for idx, val in df_mapped['MONTO'].items():
            monto_limpio = clean_monto(val)
            montos_limpios.append(monto_limpio)
            if monto_limpio is None and pd.notna(val) and str(val).strip() != '':
                montos_invalidos.append(idx + 2)
        
        df_mapped['MONTO'] = montos_limpios
        
        if montos_invalidos[:3]:
            errores.append(f"⚠️ Filas {montos_invalidos[:3]} tienen montos inválidos")
    
    if 'MONEDA' not in df_mapped.columns:
        df_mapped['MONEDA'] = 'USD'
        sugerencias.append("💡 MONEDA agregada automáticamente (USD)")
    
    # Contar filas válidas
    valid_count = 0
    for idx, row in df_mapped.iterrows():
        if (pd.notna(row.get('FECHA')) and 
            pd.notna(row.get('CLIENTE')) and 
            pd.notna(row.get('MONTO'))):
            valid_count += 1
    
    if valid_count < len(df_mapped):
        sugerencias.append(f"✅ {valid_count} filas válidas pueden importarse")
    
    # Convertir a dict para preview
    preview_data = []
    for idx, row in df_mapped.head(200).iterrows():
        row_dict = {'_original_index': int(idx)}
        for col in df_mapped.columns:
            val = row[col]
            if pd.isna(val):
                row_dict[col] = ''
            elif isinstance(val, (datetime, pd.Timestamp)):
                row_dict[col] = val.strftime('%Y-%m-%d')
            else:
                row_dict[col] = str(val)
        preview_data.append(row_dict)
    
    return {
        "status": "success" if not errores else "warning",
        "errores": errores,
        "sugerencias": sugerencias,
        "total_registros": len(df_mapped),
        "registros_validos": valid_count,
        "columnas": df_mapped.columns.tolist(),
        "columnas_originales": df.columns.tolist(),
        "column_mapping": column_mapping,
        "mapping_confidence": mapping_confidence,
        "unmapped_columns": unmapped_columns,
        "preview": preview_data,
        "estructura_detectada": "vertical" if is_vertical else "horizontal",
        "filename": file.filename,
        "sheet_name": sheet_name
    }

@router.post("/excel/confirm")
async def confirm_import(data: dict, db: Session = Depends(get_db)):
    from models.ingreso import Ingreso
    
    filename = data.get('filename', 'unknown.xlsx')
    skip_errors = data.get('skip_errors', True)  # Por defecto omitir errores
    total_rows = len(data['registros'])
    success_count = 0
    error_count = 0
    errors_log = []
    
    try:
        registros = []
        for idx, row in enumerate(data['registros']):
            try:
                row_clean = {k: v for k, v in row.items() if not k.startswith('_')}
                
                # Parsear fecha
                fecha = parse_fecha(row_clean.get('FECHA'))
                if not fecha:
                    if not skip_errors:
                        raise ValueError(f"Fecha inválida: {row_clean.get('FECHA')}")
                    error_count += 1
                    errors_log.append({"row": idx + 1, "error": "Fecha inválida"})
                    continue
                
                # Limpiar monto
                monto = clean_monto(row_clean.get('MONTO'))
                if monto is None:
                    if not skip_errors:
                        raise ValueError(f"Monto inválido: {row_clean.get('MONTO')}")
                    error_count += 1
                    errors_log.append({"row": idx + 1, "error": "Monto inválido"})
                    continue
                
                cliente = row_clean.get('CLIENTE', '').strip()
                if not cliente:
                    if not skip_errors:
                        raise ValueError("Cliente vacío")
                    error_count += 1
                    errors_log.append({"row": idx + 1, "error": "Cliente vacío"})
                    continue
                
                ing = Ingreso(
                    fecha=fecha,
                    cliente=cliente,
                    monto=monto,
                    moneda=row_clean.get('MONEDA', 'USD'),
                    awb=row_clean.get('AWB', ''),
                    descripcion=row_clean.get('DESCRIPCION', ''),
                    mes=row_clean.get('MES', fecha.strftime('%B').upper())
                )
                registros.append(ing)
                success_count += 1
            except Exception as e:
                error_count += 1
                errors_log.append({"row": idx + 1, "error": str(e)})
                if not skip_errors:
                    raise
        
        if registros:
            db.add_all(registros)
            db.commit()
        
        # Guardar historial
        history = ImportHistory(
            filename=filename,
            total_rows=total_rows,
            success_rows=success_count,
            error_rows=error_count,
            status='success' if error_count == 0 else 'partial',
            errors_log=errors_log if errors_log else None,
            column_mapping=data.get('column_mapping')
        )
        db.add(history)
        db.commit()
        
        msg = f"✅ {success_count} registros importados"
        if error_count > 0:
            msg += f" ({error_count} filas omitidas por errores)"
        
        return {
            "message": msg,
            "success": success_count,
            "errors": error_count
        }
    
    except Exception as e:
        db.rollback()
        
        history = ImportHistory(
            filename=filename,
            total_rows=total_rows,
            success_rows=0,
            error_rows=total_rows,
            status='failed',
            errors_log=[{"error": str(e)}]
        )
        db.add(history)
        db.commit()
        
        raise HTTPException(500, f"Error: {str(e)}")

@router.get("/excel/history")
async def get_import_history(db: Session = Depends(get_db)):
    history = db.query(ImportHistory).order_by(ImportHistory.created_at.desc()).limit(50).all()
    
    return [{
        "id": h.id,
        "filename": h.filename,
        "uploaded_by": h.uploaded_by,
        "total_rows": h.total_rows,
        "success_rows": h.success_rows,
        "error_rows": h.error_rows,
        "status": h.status,
        "created_at": h.created_at.isoformat(),
        "has_errors": h.errors_log is not None
    } for h in history]

@router.get("/excel/template")
async def download_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "INGRESOS"
    
    headers = ['FECHA', 'CLIENTE', 'DESCRIPCION', 'AWB', 'MONEDA', 'MONTO', 'MES']
    ws.append(headers)
    
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    example = ['07/11/2025', 'CORPORACION EJEMPLO SAC', 'Servicio export', '074 7014 1234', 'USD', 50000, 'NOVIEMBRE']
    ws.append(example)
    
    example_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    for col in range(1, 8):
        ws.cell(row=2, column=col).fill = example_fill
    
    dv_moneda = DataValidation(type="list", formula1='"USD,PEN"')
    ws.add_data_validation(dv_moneda)
    dv_moneda.add('E3:E1000')
    
    meses = "ENERO,FEBRERO,MARZO,ABRIL,MAYO,JUNIO,JULIO,AGOSTO,SETIEMBRE,OCTUBRE,NOVIEMBRE,DICIEMBRE"
    dv_mes = DataValidation(type="list", formula1=f'"{meses}"')
    ws.add_data_validation(dv_mes)
    dv_mes.add('G3:G1000')
    
    for col, width in zip(['A','B','C','D','E','F','G'], [14,40,35,20,12,16,16]):
        ws.column_dimensions[col].width = width
    
    ws_inst = wb.create_sheet("INSTRUCCIONES")
    instructions = [
        ["📋 GTL TEMPLATE - INGRESOS"],
        [""],
        ["✅ FECHA: DD/MM/YYYY (ej: 07/11/2025)"],
        ["✅ MONTO: Solo números (ej: 50000)"],
        ["✅ MONEDA: USD o PEN"],
        [""],
        ["💡 El sistema omitirá automáticamente filas con errores"],
    ]
    
    for row in instructions:
        ws_inst.append(row)
    
    ws_inst.column_dimensions['A'].width = 80
    ws_inst['A1'].font = Font(size=14, bold=True, color="1F4E78")
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename=GTL_Template_Ingresos.xlsx'}
    )

@router.post("/excel/import-all-sheets")
async def import_all_sheets(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """Importar TODAS las hojas del Excel automáticamente"""
    from models.ingreso import Ingreso
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(400, "Solo archivos Excel")
    
    contents = await file.read()
    xl = pd.ExcelFile(io.BytesIO(contents))
    
    total_imported = 0
    total_errors = 0
    sheets_processed = []
    
    for sheet_name in xl.sheet_names:
        try:
            df = pd.read_excel(io.BytesIO(contents), sheet_name=sheet_name)
            
            # Detectar estructura vertical
            is_vertical = df.iloc[0, 0] in ['FECHA', 'CLIENTE', 'MONTO', 'fecha', 'cliente', 'monto']
            if is_vertical:
                df = df.T
                df.columns = df.iloc[0]
                df = df[1:]
            
            # Mapeo de columnas
            column_mapping = {}
            for col in df.columns:
                matched, _ = fuzzy_match_column(str(col))
                if matched:
                    column_mapping[col] = matched
            
            df_mapped = df.rename(columns=column_mapping)
            
            # Verificar columnas requeridas
            if not all(col in df_mapped.columns for col in ['FECHA', 'CLIENTE', 'MONTO']):
                sheets_processed.append({
                    "sheet": sheet_name,
                    "status": "skipped",
                    "reason": "Columnas requeridas no encontradas"
                })
                continue
            
            # Agregar columna de origen
            df_mapped['ORIGEN_HOJA'] = sheet_name
            
            # Limpiar y validar
            registros = []
            errors = 0
            
            for idx, row in df_mapped.iterrows():
                try:
                    fecha = parse_fecha(row.get('FECHA'))
                    monto = clean_monto(row.get('MONTO'))
                    cliente = str(row.get('CLIENTE', '')).strip()
                    
                    if not fecha or not monto or not cliente:
                        errors += 1
                        continue
                    
                    ing = Ingreso(
                        fecha=fecha,
                        cliente=cliente,
                        monto=monto,
                        moneda=row.get('MONEDA', 'USD'),
                        awb=str(row.get('AWB', '')),
                        descripcion=f"{sheet_name} | {row.get('DESCRIPCION', '')}",
                        mes=row.get('MES', fecha.strftime('%B').upper())
                    )
                    registros.append(ing)
                except:
                    errors += 1
                    continue
            
            if registros:
                db.add_all(registros)
                total_imported += len(registros)
                total_errors += errors
                
                sheets_processed.append({
                    "sheet": sheet_name,
                    "status": "success",
                    "imported": len(registros),
                    "errors": errors
                })
            else:
                sheets_processed.append({
                    "sheet": sheet_name,
                    "status": "skipped",
                    "reason": "Sin datos válidos"
                })
                
        except Exception as e:
            sheets_processed.append({
                "sheet": sheet_name,
                "status": "error",
                "reason": str(e)
            })
    
    db.commit()
    
    # Guardar historial
    history = ImportHistory(
        filename=file.filename,
        total_rows=total_imported + total_errors,
        success_rows=total_imported,
        error_rows=total_errors,
        status='success' if total_imported > 0 else 'failed',
        errors_log=sheets_processed
    )
    db.add(history)
    db.commit()
    
    return {
        "message": f"✅ {total_imported} registros importados de {len([s for s in sheets_processed if s['status']=='success'])} hojas",
        "total_imported": total_imported,
        "total_errors": total_errors,
        "sheets": sheets_processed
    }
